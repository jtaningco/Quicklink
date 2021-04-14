from django.shortcuts import render, redirect
from django.utils.timezone import datetime, timedelta
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from .forms import PreviewOrderForm, OrderForm
from .filters import OrderFilter, PendingOrderFilter
from apps.products.models import Product, Size, Addon
from apps.orders.models import Order, ProductOrder
from apps.accounts.decorators import setup_required, allowed_users
from apps.accounts.models import *

from django.http import HttpResponse, JsonResponse
from decimal import Decimal
import json

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

from os import path
from datetime import datetime

# Create your views here.

# Create Excel Workbook
def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()

# Order Summary
@login_required(login_url='accounts:login')
@setup_required
@allowed_users(allowed_roles=[User.Types.MERCHANT, User.Types.ADMIN])
def orders(request):
    user = request.user
    data = request.GET.copy()
    orderFilter = OrderFilter(data, request=request)
    orders = orderFilter.qs

    current_time = datetime.now()
    file_date = current_time.strftime('%m%d%Y')

    shop_name = user.shop_info.slugify_name
    file_path = "static/xlsx/order_summaries/"
    file_name = "QL_" + shop_name + "_order_summary_" + str(file_date) + ".xlsx"

    if request.method == "POST":
        if 'download_summary' in request.POST:
            wb = load_workbook(file_path + file_name)
            sheet = wb.active
            sheet.title = 'Data'

            init_sheet = [
                ('Date of Delivery', 
                'Order Number', 
                'Customer Name', 
                'Order Details', 
                'Items', 
                'Total Amount', 
                'Order Status', 
                'Payment Status'),
            ]
            sheet_content = []

            for col in init_sheet:
                for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
                    for cell in row:
                        print(cell)

            wb.save(file_path + file_name)

    context = {'orders':orders, 'orderFilter':orderFilter}
    return render(request, 'orders/order_summary.html', context)

# Pending Orders
@login_required(login_url='accounts:login')
@setup_required
@allowed_users(allowed_roles=[User.Types.MERCHANT, User.Types.ADMIN])
def pendingOrders(request):
    # Get merchant user
    user = request.user

    # Get initial value for filter set (Alphabetical Order)
    productChoices = Product.objects.filter(user=user).order_by('name')
    if len(productChoices) > 0:
        product = productChoices[0]
        data = request.GET.copy()
        data.setdefault('product', product)

        # Get pending orders
        orders = Order.objects.filter(shop=user, order_status="Pending")
        
        # Get pending product orders
        productOrders = ProductOrder.objects.filter(order__in=orders).order_by('product')

        # Product filter form
        productFilter = PendingOrderFilter(data, request=request, queryset=productOrders)
        products = productFilter.qs
        sizes = Size.objects.filter(product=productFilter.get_product)
    else:
        product = None
        orders = Order.objects.filter(shop=user, order_status="Pending")
        productOrders = ProductOrder.objects.filter(order__in=orders).order_by('product')
        productFilter = PendingOrderFilter(request.GET, request=request, queryset=productOrders)
        products = productFilter.qs
        sizes = None

    try:
        addons = Addon.objects.filter(product=productFilter.get_product)
    except:
        addons = None
    
    context = {'orders':orders, 
    'products':products,
    'product':product,
    'productOrders':productOrders,
    'productFilter':productFilter,
    'sizes':sizes, 
    'addons':addons}
    return render(request, 'orders/pending_orders.html', context)

# View Products (Merchant Preview)
@login_required(login_url='accounts:login')
@setup_required
@allowed_users(allowed_roles=[User.Types.MERCHANT, User.Types.ADMIN])
def menuPreview(request):
    products = Product.objects.all()
    context = {'products':products}
    return render(request, 'orders/available_products.html', context)

# Order Form (Merchant Preview)
@login_required(login_url='accounts:login')
@setup_required
@allowed_users(allowed_roles=[User.Types.MERCHANT, User.Types.ADMIN])
def addPreviewOrder(request, pk):
    product = Product.objects.get(id=pk)
    sizes = Size.objects.filter(product=product)
    addons = Addon.objects.filter(product=product)
    form = PreviewOrderForm(initial={'product': product})
    if request.method == 'POST':
        form = PreviewOrderForm(request.POST)
        if form.is_valid():
            alert("The form has been submitted, but the data won't be portrayed in your merchant dashboard for this is merely a test.")
            return redirect('/shop/orders/')

    context = {'form':form, 'product':product, 'sizes':sizes, 'addons':addons}
    return render(request, 'orders/order_form.html', context)

# Deleting Orders (For Testing)
@login_required(login_url='accounts:login')
@setup_required
@allowed_users(allowed_roles=[User.Types.MERCHANT, User.Types.ADMIN])
def deleteOrder(request, order_pk):
    order = Order.objects.get(id=order_pk)
    if request.method == "POST":
        order.delete()
        return redirect('/shop/orders/')

    context = {'order':order}
    return render(request, 'orders/delete_order.html', context)